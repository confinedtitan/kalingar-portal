from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.db.models import Q
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter

from .models import Member, Child, Announcement, Event, Meeting, TaxMaster, MemberTax, Transaction
from accounting.models import StaffProfile
from .serializers import (
    MemberSerializer, MemberCreateSerializer, MemberUpdateSerializer,
    MemberListSerializer, ChildSerializer, UserLoginSerializer,
    PasswordChangeSerializer, AdminPasswordResetSerializer,
    AnnouncementSerializer, EventSerializer, MeetingSerializer,
    TaxMasterSerializer, MemberTaxSerializer, TransactionSerializer
)


class IsAdminOrOwner(permissions.BasePermission):
    """
    Custom permission: Admin can access everything, members can only access their own data
    """
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated
    
    def has_object_permission(self, request, view, obj):
        # Admin can access everything
        if request.user.is_staff:  # type: ignore[union-attr]
            return True
        # Members can only access their own profile
        return obj.user == request.user


class MemberViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Member CRUD operations
    """
    queryset = Member.objects.all()
    permission_classes = [IsAdminOrOwner]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]  # type: ignore[assignment]
    filterset_fields = ['is_active']
    search_fields = ['name', 'name_ta', 'phone', 'father_name', 'father_name_ta']
    ordering_fields = ['name', 'created_at', 'annual_tax', 'amount_due']
    ordering = ['name']
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'create':
            return MemberCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return MemberUpdateSerializer
        elif self.action == 'list':
            return MemberListSerializer
        return MemberSerializer
    
    def get_queryset(self):
        """Filter queryset based on user role"""
        user = self.request.user
        # IsAdminOrOwner ensures only authenticated Users reach here
        assert isinstance(user, User)
        if user.is_staff:
            # Admin sees all members
            return Member.objects.all()
        else:
            # Members see only their own profile
            return Member.objects.filter(user=user)
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        """Get current user's member profile"""
        try:
            member = Member.objects.get(user=request.user)
            serializer = MemberSerializer(member)
            return Response(serializer.data)
        except Member.DoesNotExist:
            return Response(
                {'error': 'Member profile not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get member statistics (Admin only)"""
        if not request.user.is_staff:
            return Response(
                {'error': 'Admin access required'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        total_members = Member.objects.filter(is_active=True).count()
        total_paid = Member.objects.filter(amount_due=0).count()
        total_pending = Member.objects.filter(amount_due__gt=0).count()
        
        return Response({
            'total_members': total_members,
            'members_paid': total_paid,
            'members_pending': total_pending,
        })
    
    @action(detail=True, methods=['post'])
    def add_child(self, request, pk=None):
        """Add a child to a member"""
        member = self.get_object()
        serializer = ChildSerializer(data=request.data)
        
        if serializer.is_valid():
            serializer.save(member=member)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export members to Excel (Admin only)"""
        if not request.user.is_staff:
            return Response(
                {'error': 'Admin access required'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Members"

        headers = [
            'Member ID', 'Name', 'Name (Tamil)',
            'Phone', 'DOB',
            'Address', 'Address (Tamil)',
            'Father Name', 'Father Name (Tamil)',
            'Mother Name', 'Mother Name (Tamil)',
            'Spouse Name', 'Spouse Name (Tamil)',
            'Children',
            'Annual Tax', 'Amount Paid', 'Amount Due', 'Status'
        ]

        ws.append(headers)

        # Format header row: bold and freeze pane
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = 'A2'

        members = Member.objects.prefetch_related('children').all()
        for member in members:
            children_list = '; '.join([
                f"{c.name} ({c.date_of_birth.strftime('%d/%m/%Y') if c.date_of_birth else ''}, {c.gender})"
                for c in member.children.all()
            ])
            row = [
                member.member_id,
                member.name, member.name_ta,
                member.phone,
                member.date_of_birth,
                member.address, member.address_ta,
                member.father_name, member.father_name_ta,
                member.mother_name, member.mother_name_ta,
                member.spouse_name, member.spouse_name_ta,
                children_list,
                member.annual_tax, member.amount_paid, member.amount_due,
                member.payment_status
            ]
            ws.append(row)

            # Format DOB cell (column 5) as Excel date
            dob_cell = ws.cell(row=ws.max_row, column=5)
            if dob_cell.value:
                dob_cell.number_format = 'DD/MM/YYYY'

        # Auto-size columns based on content length
        from openpyxl.utils import get_column_letter
        for col_idx, col in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                except Exception:
                    pass
            
            adjusted_width = max_length + 2
            # Set reasonable min and max bounds for column width
            adjusted_width = max(10, min(adjusted_width, 50))
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="members.xlsx"'
        wb.save(response)

        return response

    @action(detail=False, methods=['post'], url_path='import-excel',
            permission_classes=[permissions.IsAdminUser])
    def import_excel(self, request):
        """Bulk import members from a Tamil Excel file (Admin only)"""
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return Response(
                {'error': 'No file uploaded. Send a .xlsx file as "excel_file".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not excel_file.name.endswith('.xlsx'):
            return Response(
                {'error': 'Only .xlsx files are supported.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        import openpyxl
        from .utils import process_excel_workbook

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception as exc:
            return Response(
                {'error': f'Failed to read Excel file: {exc}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        results = process_excel_workbook(wb)
        return Response(results, status=status.HTTP_200_OK)


class ChildViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Child CRUD operations
    """
    queryset = Child.objects.all()
    serializer_class = ChildSerializer
    permission_classes = [IsAdminOrOwner]
    
    def get_queryset(self):
        """Filter children based on user role"""
        user = self.request.user
        assert isinstance(user, User)
        if user.is_staff:
            return Child.objects.all()
        else:
            try:
                member = Member.objects.get(user=user)
                return Child.objects.filter(member=member)
            except Member.DoesNotExist:
                return Child.objects.none()


class AuthViewSet(viewsets.ViewSet):
    """
    ViewSet for authentication operations
    """
    permission_classes = [permissions.AllowAny]
    
    @action(detail=False, methods=['post'])
    def login(self, request):
        """Login with phone and password"""
        serializer = UserLoginSerializer(data=request.data)
        
        if serializer.is_valid():
            phone = serializer.validated_data['phone']
            password = serializer.validated_data['password']
            
            # Authenticate user
            user = authenticate(username=phone, password=password)
            
            if user:
                # Narrow the type from AbstractBaseUser to User for type-checker safety.
                # authenticate() always returns a User instance when successful.
                assert isinstance(user, User)
                # Get or create token
                token, created = Token.objects.get_or_create(user=user)
                
                # Get member profile
                try:
                    member = Member.objects.get(user=user)
                    is_admin = user.is_staff
                    
                    return Response({
                        'token': token.key,
                        'user_id': user.id,  # type: ignore[attr-defined]
                        'phone': phone,
                        'is_admin': is_admin,
                        'role': 'ADMIN' if is_admin else 'MEMBER',
                        'member_id': member.id,  # type: ignore[attr-defined]
                        'name': member.name,
                        'password_reset_required': member.password_reset_required,
                    })
                except Member.DoesNotExist:
                    # Check for StaffProfile (e.g. Accountant)
                    try:
                        staff = StaffProfile.objects.get(user=user)
                        if not staff.is_active:
                            return Response(
                                {'error': 'Account is deactivated.'},
                                status=status.HTTP_403_FORBIDDEN,
                            )
                        return Response({
                            'token': token.key,
                            'user_id': user.id,  # type: ignore[attr-defined]
                            'phone': phone,
                            'is_admin': False,
                            'role': staff.role,  # 'ACCOUNTANT'
                            'staff_id': staff.id,  # type: ignore[attr-defined]
                            'name': staff.name,
                            'password_reset_required': False,
                        })
                    except StaffProfile.DoesNotExist:
                        pass

                    # Admin user without member profile
                    return Response({
                        'token': token.key,
                        'user_id': user.id,  # type: ignore[attr-defined]
                        'phone': phone,
                        'is_admin': user.is_staff,
                        'role': 'ADMIN' if user.is_staff else 'UNKNOWN',
                        'name': user.get_username(),  # get_username() is defined on AbstractBaseUser
                        'password_reset_required': False,
                    })
            else:
                return Response(
                    {'error': 'Invalid credentials'},
                    status=status.HTTP_401_UNAUTHORIZED
                )
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def logout(self, request):
        """Logout user"""
        try:
            request.user.auth_token.delete()
            return Response({'message': 'Successfully logged out'})
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def change_password(self, request):
        """Change user password"""
        serializer = PasswordChangeSerializer(data=request.data)
        
        if serializer.is_valid():
            user = request.user
            old_password = serializer.validated_data['old_password']
            new_password = serializer.validated_data['new_password']
            
            if user.check_password(old_password):
                user.set_password(new_password)
                user.save()
                
                # Clear password reset required flag for member
                try:
                    member = Member.objects.get(user=user)
                    member.password_reset_required = False
                    member.save()
                except Member.DoesNotExist:
                    pass
                
                return Response({'message': 'Password changed successfully'})
            else:
                return Response(
                    {'error': 'Old password is incorrect'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def reset_password(self, request):
        """Admin only: Reset member password to default (phone number)"""
        # Check if user is admin
        if not request.user.is_staff:
            return Response(
                {'error': 'Admin access required'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = AdminPasswordResetSerializer(data=request.data)
        
        if serializer.is_valid():
            member_id = serializer.validated_data['member_id']
            
            try:
                member = Member.objects.get(id=member_id)
                user = member.user
                
                # Reset password to phone number
                new_password = member.phone
                user.set_password(new_password)
                user.save()
                
                # Set password reset required flag
                member.password_reset_required = True
                member.save()
                
                return Response({
                    'message': f'Password reset successfully for {member.name}',
                    'temporary_password': new_password,
                    'member_id': member.id,  # type: ignore[attr-defined]
                })
            except Member.DoesNotExist:
                return Response(
                    {'error': 'Member not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class IsAdminOrReadOnly(permissions.BasePermission):
    """Admin can do anything, others can only read"""
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        # is_staff lives on User/PermissionsMixin, not AbstractBaseUser — guard with isinstance
        return isinstance(request.user, User) and request.user.is_staff


class AnnouncementViewSet(viewsets.ModelViewSet):
    serializer_class = AnnouncementSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrReadOnly]

    def get_queryset(self):
        # IsAuthenticated ensures request.user is a real User here
        assert isinstance(self.request.user, User)
        if self.request.user.is_staff:
            return Announcement.objects.all()
        return Announcement.objects.filter(is_active=True)


class EventViewSet(viewsets.ModelViewSet):
    serializer_class = EventSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrReadOnly]

    def get_queryset(self):
        # IsAuthenticated ensures request.user is a real User here
        assert isinstance(self.request.user, User)
        if self.request.user.is_staff:
            return Event.objects.all()
        return Event.objects.filter(is_active=True)


class MeetingViewSet(viewsets.ModelViewSet):
    serializer_class = MeetingSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrReadOnly]

    def get_queryset(self):
        # IsAuthenticated ensures request.user is a real User here
        assert isinstance(self.request.user, User)
        if self.request.user.is_staff:
            return Meeting.objects.all()
        return Meeting.objects.filter(is_active=True)

class TaxMasterViewSet(viewsets.ModelViewSet):
    queryset = TaxMaster.objects.all()
    serializer_class = TaxMasterSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrReadOnly]
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAdminUser])
    def generate_taxes(self, request):
        tax_id = request.data.get('tax_id')
        if not tax_id:
            return Response({'error': 'tax_id is required'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            tax_master = TaxMaster.objects.get(id=tax_id)
        except TaxMaster.DoesNotExist:
            return Response({'error': 'TaxMaster not found'}, status=status.HTTP_404_NOT_FOUND)
            
        if tax_master.status == 'Generated':
            return Response({'error': 'Taxes have already been generated for this event'}, status=status.HTTP_400_BAD_REQUEST)
            
        from decimal import Decimal
        from django.utils import timezone
        from accounting.models import AccountHead, AccountTransaction
        
        # Get or create the AccountHead 'Tax Kodai'
        account_head, _ = AccountHead.objects.get_or_create(
            name='Tax Kodai',
            defaults={
                'head_type': 'Event',
                'is_active': True,
                'created_by': request.user if request.user.is_authenticated else None
            }
        )
        
        members = Member.objects.filter(is_active=True)
        created_count = 0
        for member in members:
            tax_count = Decimal('1.0')
            for child in member.children.all():
                if child.gender == 'Male':
                    if child.marital_status == 'Unmarried':
                        tax_count += Decimal('0.5')
                    else:
                        tax_count += Decimal('1.0')
            
            total_tax = (tax_count * tax_master.base_amount).quantize(Decimal('0.00'))
            
            obj, created = MemberTax.objects.update_or_create(
                member=member,
                tax=tax_master,
                defaults={
                    'tax_count': tax_count,
                    'total_tax': total_tax,
                }
            )
            if created:
                created_count += 1
                
            # Create account debit/bill entry (Income type transaction) for the member
            if total_tax > 0:
                AccountTransaction.objects.create(
                    account_head=account_head,
                    transaction_type='INCOME',
                    amount=total_tax,
                    transaction_date=timezone.now().date(),
                    payment_mode='Credit',
                    member=member,
                    donor_name=member.name,
                    donor_contact=member.phone or '',
                    purpose=f"Tax Kodai: {tax_master.name}",
                    entered_by=request.user
                )
                
        # Update tax event status and generated date
        tax_master.status = 'Generated'
        tax_master.generated_date = timezone.now().date()
        tax_master.save()
        
        return Response({'message': f'Generated {created_count} new taxes for {tax_master.name}'})

class MemberTaxViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = MemberTaxSerializer
    permission_classes = [IsAdminOrOwner]
    
    def get_queryset(self):
        user = self.request.user
        assert isinstance(user, User)
        if user.is_staff:
            return MemberTax.objects.all()
        try:
            member = Member.objects.get(user=user)
            return MemberTax.objects.filter(member=member)
        except Member.DoesNotExist:
            return MemberTax.objects.none()

class TransactionViewSet(viewsets.ModelViewSet):
    serializer_class = TransactionSerializer
    permission_classes = [IsAdminOrOwner]
    
    def get_queryset(self):
        user = self.request.user
        assert isinstance(user, User)
        if user.is_staff:
            return Transaction.objects.all()
        try:
            member = Member.objects.get(user=user)
            return Transaction.objects.filter(member=member)
        except Member.DoesNotExist:
            return Transaction.objects.none()
            
    def perform_create(self, serializer):
        user = self.request.user
        assert isinstance(user, User)
        if user.is_staff and 'member' in serializer.validated_data:
            member = serializer.validated_data['member']
        else:
            member = Member.objects.get(user=user)
            
        transaction = serializer.save(member=member)
        
        if transaction.member_tax and transaction.transaction_type == 'Payment':
            member_tax = transaction.member_tax
            member_tax.amount_paid += transaction.amount
            member_tax.save()
