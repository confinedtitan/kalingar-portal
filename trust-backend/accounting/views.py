"""
DRF ViewSets for the accounting app.

Includes:
- StaffProfileViewSet  (Admin-only CRUD for accountant accounts)
- AccountHeadViewSet    (Accountant + Admin)
- AccountTransactionViewSet (with soft-delete, audit log, filters)
- ReceiptViewSet        (read-only + download)
- Excel export endpoints
- Summary endpoints
"""

import json
from decimal import Decimal

from django.contrib.auth.models import User
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter

from .models import StaffProfile, AccountHead, AccountTransaction, Receipt, TrustAccount
from .serializers import (
    StaffProfileSerializer, StaffProfileCreateSerializer,
    AccountHeadSerializer,
    AccountTransactionSerializer, AccountTransactionCreateSerializer,
    ReceiptSerializer, TrustAccountSerializer,
)
from .permissions import IsAccountantOrAdmin, IsAdmin


# ---------------------------------------------------------------------------
# StaffProfile — Admin-only management of accountant accounts
# ---------------------------------------------------------------------------

class StaffProfileViewSet(viewsets.ModelViewSet):
    """
    Admin-only CRUD for staff (accountant) profiles.
    """
    queryset = StaffProfile.objects.select_related('user').all()
    permission_classes = [permissions.IsAuthenticated, IsAdmin]

    def get_serializer_class(self):
        if self.action == 'create':
            return StaffProfileCreateSerializer
        return StaffProfileSerializer

    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        """Deactivate a staff profile."""
        profile = self.get_object()
        profile.is_active = False
        profile.save()
        return Response({'message': f'{profile.name} deactivated.'})

    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """Re-activate a staff profile."""
        profile = self.get_object()
        profile.is_active = True
        profile.save()
        return Response({'message': f'{profile.name} activated.'})


# ---------------------------------------------------------------------------
# TrustAccount
# ---------------------------------------------------------------------------

class TrustAccountViewSet(viewsets.ModelViewSet):
    """
    Trust Accounts management ViewSet.
    """
    queryset = TrustAccount.objects.all()
    serializer_class = TrustAccountSerializer
    permission_classes = [permissions.IsAuthenticated, IsAccountantOrAdmin]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['account_type', 'status', 'associated_entity_type']
    search_fields = ['account_name', 'account_number', 'bank_name']
    ordering_fields = ['account_name', 'created_date']

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        """Deactivate a trust account."""
        account = self.get_object()
        account.status = 'Inactive'
        account.deactivated_date = timezone.now()
        account.save()
        return Response({'message': f'"{account.account_name}" deactivated.', 'deactivated_date': account.deactivated_date})

    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """Re-activate a trust account."""
        account = self.get_object()
        account.status = 'Active'
        account.deactivated_date = None
        account.save()
        return Response({'message': f'"{account.account_name}" activated.'})

    @action(detail=True, methods=['get'])
    def ledger(self, request, pk=None):
        """Ledger statement: Chronological list of transactions and running balance."""
        account = self.get_object()
        txns = AccountTransaction.objects.filter(
            trust_account=account, is_deleted=False
        ).order_by('transaction_date', 'id')
        
        data = []
        balance = Decimal('0')
        for t in txns:
            if t.transaction_type == 'DEBIT':
                balance += t.amount
            else:
                balance -= t.amount
                
            data.append({
                'id': t.id,
                'transaction_date': str(t.transaction_date),
                'transaction_type': t.transaction_type,
                'amount': str(t.amount),
                'payment_mode': t.payment_mode,
                'commodity_type': t.commodity_type,
                'donor_name': t.donor_name or (t.member.name if t.member else ''),
                'donor_name_ta': t.donor_name_ta or (t.member.name_ta if t.member else ''),
                'paid_to': t.paid_to,
                'purpose': t.purpose or t.purpose_description,
                'receipt_number': t.receipt.receipt_number if hasattr(t, 'receipt') and t.receipt else None,
                'running_balance': str(balance)
            })
        return Response({
            'account_name': account.account_name,
            'account_type': account.account_type,
            'transactions': data
        })


# ---------------------------------------------------------------------------
# AccountHead
# ---------------------------------------------------------------------------

class AccountHeadViewSet(viewsets.ModelViewSet):
    """
    Account Heads — both Accountant and Admin can list/create.
    Only Admin can edit/deactivate.
    """
    serializer_class = AccountHeadSerializer
    permission_classes = [permissions.IsAuthenticated, IsAccountantOrAdmin]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active', 'head_type']
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['-created_at']

    def get_queryset(self):
        return AccountHead.objects.all()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def update(self, request, *args, **kwargs):
        # Only Admin can edit
        if not request.user.is_staff:
            return Response(
                {'error': 'Only admin can edit Account Heads.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'error': 'Only admin can edit Account Heads.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().partial_update(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        """Deactivate an account head (Admin only)."""
        if not request.user.is_staff:
            return Response(
                {'error': 'Only admin can deactivate Account Heads.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        head = self.get_object()
        head.is_active = False
        head.save()
        return Response({'message': f'"{head.name}" deactivated.'})

    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        """Per-head summary: total debits, credits, net, count."""
        head = self.get_object()
        txns = AccountTransaction.objects.filter(
            account_head=head, is_deleted=False,
        )
        total_debits = txns.filter(
            transaction_type='DEBIT',
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_credits = txns.filter(
            transaction_type='CREDIT',
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        if head.account_type in ['Asset', 'Expense']:
            net_balance = total_debits - total_credits
        else:
            net_balance = total_credits - total_debits

        return Response({
            'account_head': head.name,
            'total_income': str(total_credits),
            'total_expense': str(total_debits),
            'total_debits': str(total_debits),
            'total_credits': str(total_credits),
            'net_balance': str(net_balance),
            'transaction_count': txns.count(),
        })

    @action(detail=True, methods=['get'])
    def export(self, request, pk=None):
        """
        Export a single Account Head's transactions as .xlsx.
        Supports ?from=YYYY-MM-DD&to=YYYY-MM-DD query params.
        """
        head = self.get_object()
        return _build_export_response(
            request, [head],
            filename=f"AccountHead_{head.name.replace(' ', '_')}.xlsx",
        )

    @action(detail=False, methods=['get'], url_path='export-all')
    def export_all(self, request):
        """Export ALL account heads (Admin only)."""
        if not request.user.is_staff:
            return Response(
                {'error': 'Admin access required.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        heads = AccountHead.objects.all()
        return _build_export_response(
            request, heads, filename="All_Account_Heads.xlsx",
        )


# ---------------------------------------------------------------------------
# AccountTransaction
# ---------------------------------------------------------------------------

class AccountTransactionViewSet(viewsets.ModelViewSet):
    """
    Accountant: create / edit own / soft-delete own.
    Admin: view all / edit all / soft-delete all.
    Member: read-only access to own donations via my-donations action.
    """
    permission_classes = [permissions.IsAuthenticated, IsAccountantOrAdmin]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['account_head', 'transaction_type', 'payment_mode']
    search_fields = ['donor_name', 'paid_to', 'donor_contact', 'purpose']
    ordering_fields = ['transaction_date', 'amount', 'created_at']
    ordering = ['-transaction_date']

    def get_permissions(self):
        """
        my-donations is the only action accessible to regular Members.
        All other actions (list, retrieve, create, update, destroy) still
        require IsAccountantOrAdmin.
        """
        if self.action == 'my_donations':
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), IsAccountantOrAdmin()]

    def get_serializer_class(self):
        if self.action in ('create',):
            return AccountTransactionCreateSerializer
        return AccountTransactionSerializer

    def get_queryset(self):
        qs = AccountTransaction.objects.select_related(
            'account_head', 'entered_by', 'member', 'receipt', 'trust_account'
        )
        # By default exclude soft-deleted; admin can include them
        include_deleted = self.request.query_params.get('include_deleted')
        if not (include_deleted == 'true' and self.request.user.is_staff):
            qs = qs.filter(is_deleted=False)

        # Date range filters
        date_from = self.request.query_params.get('from')
        date_to = self.request.query_params.get('to')
        if date_from:
            qs = qs.filter(transaction_date__gte=date_from)
        if date_to:
            qs = qs.filter(transaction_date__lte=date_to)

        # Trust Account filter
        trust_acct = self.request.query_params.get('trust_account')
        if trust_acct:
            qs = qs.filter(trust_account_id=trust_acct)

        # Head type multi-filter (comma-separated, e.g. General,Kodai)
        head_types = self.request.query_params.get('head_types')
        if head_types:
            types = [t.strip() for t in head_types.split(',') if t.strip()]
            qs = qs.filter(account_head__head_type__in=types)

        # Account type multi-filter (comma-separated, e.g. Revenue,Expense)
        account_types = self.request.query_params.get('account_types')
        if account_types:
            types = [t.strip() for t in account_types.split(',') if t.strip()]
            qs = qs.filter(account_head__account_type__in=types)

        # Bank ledger filter
        is_bank = self.request.query_params.get('is_bank')
        if is_bank == 'true':
            qs = qs.filter(trust_account__account_type='Bank')

        # Cash ledger filter
        is_cash = self.request.query_params.get('is_cash')
        if is_cash == 'true':
            qs = qs.filter(trust_account__account_type='Cash')

        return qs

    def perform_create(self, serializer):
        serializer.save(entered_by=self.request.user)

    def perform_update(self, serializer):
        """Log changes before saving (audit trail)."""
        instance = self.get_object()
        user = self.request.user

        # Accountant can only edit own entries
        if not user.is_staff and instance.entered_by != user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only edit your own entries.")

        # Build change log entries
        changes = []
        for field, new_value in serializer.validated_data.items():
            old_value = getattr(instance, field, None)
            # Handle ForeignKey — compare IDs
            if hasattr(old_value, 'pk'):
                old_value = old_value.pk
            if hasattr(new_value, 'pk'):
                new_value = new_value.pk
            if str(old_value) != str(new_value):
                changes.append({
                    'field': field,
                    'old': str(old_value),
                    'new': str(new_value),
                    'at': timezone.now().isoformat(),
                    'by': user.get_username(),
                })

        if changes:
            log = list(instance.change_log or [])
            log.extend(changes)
            serializer.validated_data['change_log'] = log

        serializer.save()

    def destroy(self, request, *args, **kwargs):
        """Soft-delete instead of hard-delete."""
        instance = self.get_object()

        # Accountant can only soft-delete own entries
        if not request.user.is_staff and instance.entered_by != request.user:
            return Response(
                {'error': 'You can only delete your own entries.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        instance.is_deleted = True
        log = list(instance.change_log or [])
        log.append({
            'field': 'is_deleted',
            'old': 'False',
            'new': 'True',
            'at': timezone.now().isoformat(),
            'by': request.user.get_username(),
        })
        instance.change_log = log
        instance.save()
        return Response(
            {'message': 'Transaction soft-deleted.'},
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Overall summary across all heads."""
        qs = AccountTransaction.objects.filter(is_deleted=False)
        total_debits = qs.filter(
            transaction_type='DEBIT',
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')
        total_credits = qs.filter(
            transaction_type='CREDIT',
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        return Response({
            'total_income': str(total_credits),
            'total_expense': str(total_debits),
            'net_balance': str(total_credits - total_debits),
            'total_transactions': qs.count(),
        })

    @action(detail=False, methods=['get'], url_path='commodity-summary')
    def commodity_summary(self, request):
        """
        Natively calculates and summarizes physical weight assets: Gold and Silver,
        alongside fiat currency.
        """
        qs = AccountTransaction.objects.filter(is_deleted=False)
        
        # Fiat summary (exclude Commodities payment mode)
        fiat_credits = qs.filter(transaction_type='CREDIT').exclude(payment_mode='Commodities').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        fiat_debits = qs.filter(transaction_type='DEBIT').exclude(payment_mode='Commodities').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Gold physical weight summary (payment_mode='Commodities', commodity_type='Gold')
        gold_credits = qs.filter(payment_mode='Commodities', commodity_type='Gold', transaction_type='CREDIT').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        gold_debits = qs.filter(payment_mode='Commodities', commodity_type='Gold', transaction_type='DEBIT').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Silver physical weight summary (payment_mode='Commodities', commodity_type='Silver')
        silver_credits = qs.filter(payment_mode='Commodities', commodity_type='Silver', transaction_type='CREDIT').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        silver_debits = qs.filter(payment_mode='Commodities', commodity_type='Silver', transaction_type='DEBIT').aggregate(total=Sum('amount'))['total'] or Decimal('0')

        # Other physical weight summary
        other_credits = qs.filter(payment_mode='Commodities', commodity_type='Other', transaction_type='CREDIT').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        other_debits = qs.filter(payment_mode='Commodities', commodity_type='Other', transaction_type='DEBIT').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        return Response({
            'fiat': {
                'total_credits': str(fiat_credits),
                'total_debits': str(fiat_debits),
                'net_balance': str(fiat_credits - fiat_debits),
            },
            'gold': {
                'total_credits': str(gold_credits),
                'total_debits': str(gold_debits),
                'net_balance': str(gold_credits - gold_debits),
            },
            'silver': {
                'total_credits': str(silver_credits),
                'total_debits': str(silver_debits),
                'net_balance': str(silver_credits - silver_debits),
            },
            'other': {
                'total_credits': str(other_credits),
                'total_debits': str(other_debits),
                'net_balance': str(other_credits - other_debits),
            }
        })

    @action(detail=False, methods=['get'], url_path='member-balances')
    def member_balances(self, request):
        """
        Roster of all members showing their Name and their exact running financial balance.
        Balance = Cumulative DEBIT (due/billed) - Cumulative CREDIT (paid).
        """
        from members.models import Member
        members = Member.objects.filter(is_active=True, is_family_head=True).order_by('name')
        
        data = []
        for m in members:
            debits = AccountTransaction.objects.filter(
                member=m, transaction_type='DEBIT', is_deleted=False
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            
            credits = AccountTransaction.objects.filter(
                member=m, transaction_type='CREDIT', is_deleted=False
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            
            data.append({
                'id': m.id,
                'member_id': m.member_id,
                'name': m.name,
                'name_ta': m.name_ta,
                'phone': m.phone,
                'total_debits': str(debits),
                'total_credits': str(credits),
                'running_balance': str(debits - credits)
            })
        return Response(data)

    @action(detail=False, methods=['get'], url_path='my-donations')
    def my_donations(self, request):
        """
        Member-facing endpoint: returns all INCOME transactions linked
        to the logged-in member.
        """
        user = request.user
        try:
            member = user.member_profile
        except Exception:
            return Response(
                {'error': 'No member profile found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        qs = AccountTransaction.objects.filter(
            transaction_type='INCOME',
            member=member,
            is_deleted=False,
        ).select_related('account_head', 'receipt').order_by('-transaction_date')

        serializer = AccountTransactionSerializer(qs, many=True)
        # Also compute lifetime total and per-head breakdown
        lifetime_total = qs.aggregate(t=Sum('amount'))['t'] or Decimal('0')
        per_head = (
            qs.values('account_head__name')
            .annotate(total=Sum('amount'), count=Count('id'))
            .order_by('-total')
        )

        return Response({
            'donations': serializer.data,
            'lifetime_total': str(lifetime_total),
            'per_head_breakdown': list(per_head),
        })


# ---------------------------------------------------------------------------
# Receipt
# ---------------------------------------------------------------------------

class ReceiptViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only viewset for receipts + PDF download.

    Queryset scoping rules:
    - Admin (is_staff=True)  → all receipts.
    - Accountant (StaffProfile) → all receipts.
    - Member                  → only receipts whose transaction.member
                                 matches the logged-in user's member profile.
    - Anyone else              → empty queryset.
    """

    serializer_class = ReceiptSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        base_qs = Receipt.objects.select_related(
            'transaction', 'transaction__account_head',
            'transaction__entered_by', 'transaction__member',
        )

        # Admin: unrestricted
        if user.is_staff:
            return base_qs

        # Accountant: unrestricted
        try:
            profile = user.staff_profile
            if profile.role == 'ACCOUNTANT' and profile.is_active:
                return base_qs
        except Exception:
            pass

        # Member: scope to own receipts only
        try:
            member = user.member_profile
            return base_qs.filter(transaction__member=member)
        except Exception:
            pass

        # Unknown user type — return nothing
        return base_qs.none()

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """Download receipt PDF."""
        receipt = self.get_object()  # enforces get_queryset() scoping

        # If PDF doesn't exist yet, generate it on the fly
        if not receipt.pdf_file:
            from .receipt_generator import generate_receipt_pdf
            from django.core.files.base import ContentFile

            pdf_bytes = generate_receipt_pdf(receipt.transaction)
            filename = f"{receipt.receipt_number.replace('/', '_')}.pdf"
            receipt.pdf_file.save(filename, ContentFile(pdf_bytes), save=True)

        response = HttpResponse(
            receipt.pdf_file.read(),
            content_type='application/pdf',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="{receipt.receipt_number.replace("/", "_")}.pdf"'
        )
        return response


# ---------------------------------------------------------------------------
# Excel export helper (openpyxl)
# ---------------------------------------------------------------------------

def _build_export_response(request, account_heads, filename):
    """
    Build an .xlsx HttpResponse with Income / Expense / Summary sheets
    for the given account head(s).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    date_from = request.query_params.get('from')
    date_to = request.query_params.get('to')

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4338CA', end_color='4338CA', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    overall_debits = Decimal('0')
    overall_credits = Decimal('0')

    for head in account_heads:
        qs = AccountTransaction.objects.filter(
            account_head=head, is_deleted=False,
        ).select_related('member', 'entered_by', 'receipt')

        if date_from:
            qs = qs.filter(transaction_date__gte=date_from)
        if date_to:
            qs = qs.filter(transaction_date__lte=date_to)

        income_txns = qs.filter(transaction_type='CREDIT').order_by('transaction_date')
        expense_txns = qs.filter(transaction_type='DEBIT').order_by('transaction_date')

        head_income = income_txns.aggregate(t=Sum('amount'))['t'] or Decimal('0')
        head_expense = expense_txns.aggregate(t=Sum('amount'))['t'] or Decimal('0')
        overall_credits += head_income
        overall_debits += head_expense

        # Truncate sheet name to 31 chars (Excel limit)
        safe_name = head.name[:28]

        # --- Income (Credit) sheet ---
        ws_income = wb.create_sheet(title=f"{safe_name} Credits")
        ws_income.cell(row=1, column=1, value='ஸ்ரீ ராமஜெயம் | Sri Ramajeyam').font = Font(bold=True, size=14, color='D97706')
        ws_income.merge_cells('A1:L1')
        ws_income.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        income_headers = [
            'Date', 'Donor Name', 'Donor Name (Tamil)', 'Contact', 'Member ID', 'Member Name',
            'Amount', 'Payment Mode', 'Purpose/Remarks', 'Purpose/Remarks (Tamil)', 'Receipt No', 'Entered By',
        ]
        for col, hdr in enumerate(income_headers, 1):
            cell = ws_income.cell(row=2, column=col, value=hdr)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, txn in enumerate(income_txns, 3):
            entered_name = _get_display_name(txn.entered_by)
            receipt_no = ''
            try:
                receipt_no = txn.receipt.receipt_number
            except Exception:
                pass

            vals = [
                txn.transaction_date.strftime('%Y-%m-%d'),
                txn.donor_name,
                txn.donor_name_ta,
                txn.donor_contact,
                txn.member.member_id if txn.member else '',
                txn.member.name if txn.member else '',
                float(txn.amount),
                txn.payment_mode,
                txn.purpose,
                txn.purpose_ta,
                receipt_no,
                entered_name,
            ]
            for col, val in enumerate(vals, 1):
                cell = ws_income.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

        # --- Expense (Debit) sheet ---
        ws_expense = wb.create_sheet(title=f"{safe_name} Debits")
        ws_expense.cell(row=1, column=1, value='ஸ்ரீ ராமஜெயம் | Sri Ramajeyam').font = Font(bold=True, size=14, color='D97706')
        ws_expense.merge_cells('A1:K1')
        ws_expense.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        expense_headers = [
            'Date', 'Paid To', 'Paid To (Tamil)', 'Purpose/Description', 'Purpose/Description (Tamil)',
            'Amount', 'Payment Mode', 'Bill Reference', 'Receipt No', 'Entered By',
        ]
        for col, hdr in enumerate(expense_headers, 1):
            cell = ws_expense.cell(row=2, column=col, value=hdr)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, txn in enumerate(expense_txns, 3):
            entered_name = _get_display_name(txn.entered_by)
            receipt_no = ''
            try:
                receipt_no = txn.receipt.receipt_number
            except Exception:
                pass

            vals = [
                txn.transaction_date.strftime('%Y-%m-%d'),
                txn.paid_to,
                txn.paid_to_ta,
                txn.purpose_description,
                txn.purpose_description_ta,
                float(txn.amount),
                txn.payment_mode,
                txn.bill_reference,
                receipt_no,
                entered_name,
            ]
            for col, val in enumerate(vals, 1):
                cell = ws_expense.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

    # --- Summary sheet ---
    ws_summary = wb.create_sheet(title='Summary', index=0)
    summary_header_fill = PatternFill(
        start_color='059669', end_color='059669', fill_type='solid',
    )

    ws_summary.cell(row=1, column=1, value='ஸ்ரீ ராமஜெயம் | Sri Ramajeyam').font = Font(
        bold=True, size=14, color='D97706',
    )
    ws_summary.merge_cells('A1:E1')
    ws_summary.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    ws_summary.cell(row=2, column=1, value='Account Head Report Summary').font = Font(
        bold=True, size=14,
    )
    ws_summary.merge_cells('A2:E2')

    date_range_str = ''
    if date_from or date_to:
        date_range_str = f"From: {date_from or 'Start'} — To: {date_to or 'Present'}"
        ws_summary.cell(row=3, column=1, value=date_range_str)
        ws_summary.merge_cells('A3:E3')

    start_row = 5
    for col, hdr in enumerate(
        ['Account Head', 'Account Type', 'Total Credits', 'Total Debits', 'Net Balance'], 1,
    ):
        cell = ws_summary.cell(row=start_row, column=col, value=hdr)
        cell.font = header_font
        cell.fill = summary_header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row = start_row + 1
    overall_net_balance = Decimal('0')
    for head in account_heads:
        qs = AccountTransaction.objects.filter(
            account_head=head, is_deleted=False,
        )
        if date_from:
            qs = qs.filter(transaction_date__gte=date_from)
        if date_to:
            qs = qs.filter(transaction_date__lte=date_to)

        hi = qs.filter(transaction_type='CREDIT').aggregate(t=Sum('amount'))['t'] or Decimal('0')
        he = qs.filter(transaction_type='DEBIT').aggregate(t=Sum('amount'))['t'] or Decimal('0')

        if head.account_type in ['Asset', 'Expense']:
            net_balance = he - hi
        else:
            net_balance = hi - he
            
        overall_net_balance += net_balance

        ws_summary.cell(row=row, column=1, value=head.name).border = thin_border
        ws_summary.cell(row=row, column=2, value=head.account_type).border = thin_border
        ws_summary.cell(row=row, column=3, value=float(hi)).border = thin_border
        ws_summary.cell(row=row, column=4, value=float(he)).border = thin_border
        ws_summary.cell(row=row, column=5, value=float(net_balance)).border = thin_border
        row += 1

    # Totals row
    total_font = Font(bold=True, size=11)
    ws_summary.cell(row=row, column=1, value='TOTAL').font = total_font
    ws_summary.cell(row=row, column=2, value='').border = thin_border
    ws_summary.cell(
        row=row, column=3, value=float(overall_credits),
    ).font = total_font
    ws_summary.cell(
        row=row, column=4, value=float(overall_debits),
    ).font = total_font
    ws_summary.cell(
        row=row, column=5, value=float(overall_net_balance),
    ).font = total_font
    for col in range(1, 6):
        ws_summary.cell(row=row, column=col).border = thin_border

    # Auto-width columns
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    col_letter = cell.column_letter
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Build response
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _get_display_name(user):
    """Get a human-readable name for a user."""
    try:
        return user.staff_profile.name
    except Exception:
        pass
    try:
        return user.member_profile.name
    except Exception:
        pass
    return user.get_username()


from rest_framework.views import APIView
import glob
import os
from django.conf import settings

class SystemSettingsView(APIView):
    """
    API endpoint to get and set system-level receipt templates.
    """
    permission_classes = [permissions.IsAuthenticated, IsAccountantOrAdmin]

    def get(self, request):
        # Load active settings
        settings_path = os.path.join(settings.BASE_DIR, 'accounting', 'system_settings.json')
        active_template = "tax_receipt_template.html"
        if os.path.exists(settings_path):
            try:
                with open(settings_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    active_template = data.get('active_receipt_template', 'tax_receipt_template.html')
            except Exception:
                pass

        # Scan folder for HTML files
        dir_path = os.path.join(settings.BASE_DIR, 'accounting')
        html_files = glob.glob(os.path.join(dir_path, '*.html'))
        templates = [os.path.basename(f) for f in html_files]
        if "tax_receipt_template.html" not in templates:
            templates.append("tax_receipt_template.html")

        # Get content of the active template
        template_content = ""
        active_template_path = os.path.join(dir_path, active_template)
        if os.path.exists(active_template_path):
            try:
                with open(active_template_path, 'r', encoding='utf-8') as f:
                    template_content = f.read()
            except Exception:
                pass

        return Response({
            'active_receipt_template': active_template,
            'templates': list(set(templates)),
            'template_content': template_content,
        })

    def post(self, request):
        action = request.data.get('action')
        dir_path = os.path.join(settings.BASE_DIR, 'accounting')
        
        if action == 'save_template_content':
            template_name = request.data.get('template_name')
            content = request.data.get('content')
            if not template_name or not content:
                return Response({'error': 'Missing template_name or content'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Clean filename to prevent path traversal
            template_name = os.path.basename(template_name)
            if not template_name.endswith('.html'):
                template_name += '.html'
            
            target_path = os.path.join(dir_path, template_name)
            try:
                with open(target_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                return Response({'message': f'Template "{template_name}" saved successfully.'})
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        else:
            # Standard settings update: active template
            active_template = request.data.get('active_receipt_template')
            if not active_template:
                return Response({'error': 'Missing active_receipt_template'}, status=status.HTTP_400_BAD_REQUEST)

            active_template = os.path.basename(active_template)
            settings_path = os.path.join(dir_path, 'system_settings.json')
            
            # Check if template actually exists
            if not os.path.exists(os.path.join(dir_path, active_template)):
                return Response({'error': 'Template file does not exist'}, status=status.HTTP_404_NOT_FOUND)

            try:
                config_data = {}
                if os.path.exists(settings_path):
                    with open(settings_path, 'r', encoding='utf-8') as f:
                        config_data = json.load(f)
                config_data['active_receipt_template'] = active_template
                with open(settings_path, 'w', encoding='utf-8') as f:
                    json.dump(config_data, f, indent=4)
                return Response({'message': 'Active receipt template updated successfully.'})
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
